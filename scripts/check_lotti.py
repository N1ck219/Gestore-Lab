import pandas as pd
from openpyxl import load_workbook

file_path = r'database\Materie Prime.xlsm'
lotto_id = '2260146'
lotto_id2 = 'UDN/075-26'

print("=" * 60)
print("TEST 1: Lettura con pandas (default)")
print("=" * 60)
df = pd.read_excel(file_path, sheet_name='Lotti interni', engine='openpyxl')
row = df[df['LOTTO INTERNO'].astype(str).str.strip() == lotto_id]
if not row.empty:
    r = row.iloc[0]
    print(f"  Lotto: {lotto_id}")
    print(f"  Qnt.Arrivata = {r.get('Qnt. Arrivata')}")
    print(f"  Consumi      = {r.get('Consumi')}")
    print(f"  Giacenza     = {r.get('Giacenza')}")
    print(f"  Etich.       = {r.get('Etich.')}")

row2 = df[df['LOTTO INTERNO'].astype(str).str.strip() == lotto_id2]
if not row2.empty:
    r2 = row2.iloc[0]
    print(f"\n  Lotto: {lotto_id2}")
    print(f"  Qnt.Arrivata = {r2.get('Qnt. Arrivata')}")
    print(f"  Consumi      = {r2.get('Consumi')}")
    print(f"  Giacenza     = {r2.get('Giacenza')}")
    print(f"  Etich.       = {r2.get('Etich.')}")

print("\n" + "=" * 60)
print("TEST 2: Lettura con openpyxl data_only=True")
print("=" * 60)
wb = load_workbook(file_path, data_only=True)
ws = wb['Lotti interni']

# Leggiamo gli header dalla prima riga
headers = [cell.value for cell in ws[1]]
print(f"Headers: {headers[:25]}")

# Cerchiamo le righe con i lotti specifici
lotto_col_idx = headers.index('LOTTO INTERNO')
consumi_idx = headers.index('Consumi') if 'Consumi' in headers else None
giacenza_idx = headers.index('Giacenza') if 'Giacenza' in headers else None
etich_idx = headers.index('Etich.') if 'Etich.' in headers else None

print(f"\nIndici colonne: Consumi={consumi_idx}, Giacenza={giacenza_idx}, Etich={etich_idx}")

for row in ws.iter_rows(min_row=2, values_only=True):
    lotto_val = str(row[lotto_col_idx]).strip() if row[lotto_col_idx] else ''
    if lotto_val in [lotto_id, lotto_id2]:
        print(f"\n  Lotto: {lotto_val}")
        print(f"  Consumi  (col {consumi_idx}) = {row[consumi_idx]}")
        print(f"  Giacenza (col {giacenza_idx}) = {row[giacenza_idx]}")
        print(f"  Etich    (col {etich_idx}) = {row[etich_idx]}")

print("\n" + "=" * 60)
print("TEST 3: Lettura con openpyxl data_only=False (formule)")
print("=" * 60)
wb2 = load_workbook(file_path, data_only=False)
ws2 = wb2['Lotti interni']

for row in ws2.iter_rows(min_row=2, values_only=True):
    lotto_val = str(row[lotto_col_idx]).strip() if row[lotto_col_idx] else ''
    if lotto_val in [lotto_id, lotto_id2]:
        print(f"\n  Lotto: {lotto_val}")
        print(f"  Consumi  = {row[consumi_idx]}  (formula?)")
        print(f"  Giacenza = {row[giacenza_idx]}  (formula?)")
        print(f"  Etich    = {row[etich_idx]}  (formula?)")

wb.close()
wb2.close()
